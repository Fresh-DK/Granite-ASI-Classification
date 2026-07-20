from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_workbook(path: Path | str) -> None:
    """Apply consistent, readable formatting to a pandas-generated workbook."""
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = True
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 30

        sampled_last_row = min(sheet.max_row, 250)
        for column_index in range(1, sheet.max_column + 1):
            values = [sheet.cell(row, column_index).value for row in range(1, sampled_last_row + 1)]
            max_length = max((len(str(value)) for value in values if value is not None), default=0)
            header = str(sheet.cell(1, column_index).value or "")
            if max_length > 28 or len(header) > 22:
                width = min(max(max_length + 2, 16), 38)
            else:
                width = min(max(max_length + 2, 10), 24)
            sheet.column_dimensions[get_column_letter(column_index)].width = width

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
                cell.alignment = Alignment(vertical="top", wrap_text=False)
    workbook.save(path)
